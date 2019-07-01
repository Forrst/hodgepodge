#!/usr/bin/python
# -*- coding:utf-8 -*- 
'''
作者:jia.zhou@aliyun.com
创建时间:2019-05-24 上午10:44
'''
#! flask/bin/python

from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from openpyxl import Workbook
from openpyxl.styles import numbers

from openpyxl.chart import (
    LineChart,
    BarChart,
    Reference,
    Series,
)

#engine = create_engine('mysql+pymysql://root:zunjiazichan123@192.168.2.231:3306/nginxlog')
#conn = engine.connect()

#Session=sessionmaker(bind=engine)
#dbsession=Session()

#sqlstr = """ SELECT
#  FROM_UNIXTIME(UNIX_TIMESTAMP(TIME), '%Y%m%d') AS statday,
#  SUM(COUNT),
#  SUM(serviceresponsetime)
#FROM
#  response
#WHERE TIME >= "2019-05-01 00:00:00"
#  GROUP BY statday   """

#row = dbsession.execute(sqlstr)

#print(row.fetchone())from openpyxl.styles import numbers
#print(row.fetchmany())
#print(row.fetchall())


rows = [
    ('aa', 10, 130),
    ('bb', 40, 160),
    ('cc', 50, 170),
    ('dd', 20, 110),
    ('ee', 10, 140),
    ('ff', 50, 130),
]


wb = Workbook()
#sheet = wb.create_sheet()
sheet = wb.get_active_sheet()
sheet.title = u'服务质量日统计表'

sheet['A1'] =u'日期'
sheet['B1'] =u'访问次数'
sheet['C1'] =u'访问总时长'
sheet['D1'] =u'访问平均时长'

for record in rows:
    linedata = list(record)
    print('===',linedata)
    #if linedata[1] != 0 :
    #  linedata.append(linedata[2]/linedata[1])
    #else:
    #  linedata.append('#')

    sheet.append(linedata)

sheet.column_dimensions['A'].width = 12
sheet.column_dimensions['B'].width = 15
sheet.column_dimensions['C'].width = 0
sheet.column_dimensions['D'].width = 15

for i in range(2,sheet.max_row+1):
    sheet.cell(row=i,column=4).number_format = numbers.FORMAT_NUMBER_00

# Create first chart
c1 = BarChart()
# 垂直条  bar为横向图
c1.type = "col"
c1.style = 10
c1.shape = 4

# c1.x_axis.title = u'日期'
c1.x_axis.title = ['aa','bb','cc','dd','ee','ff']
c1.y_axis.title = u'访问量'
c1.y_axis.majorGridlines = None
c1.title = u'访问量日统计'

data = Reference(sheet, min_col=2, max_col=3, min_row=2,max_row=7)
cats = Reference(sheet, min_col=1, min_row=2, max_row=7)

c1.add_data(data, titles_from_data=True, from_rows=True)
c1.set_categories(cats)


# Create a second chart
c2 = LineChart()
v2 = Reference(sheet, min_col=1, min_row=2, max_col=4)
c2.add_data(v2, titles_from_data=True, from_rows=True)
c2.y_axis.axId = 200
c2.y_axis.title = u"网络流量"

# Display y-axis of the second chart on the right by setting it to cross the x-axis at its maximum
# c1.y_axis.crosses = "max"
# c1 += c2
sheet.add_image()
sheet.add_chart(c2, "F6")

# wb.save(r'd:\\test.xlsx')

wb.save(r'/home/eos/tests.xlsx')

#conn.close()

# import matplotlib.pyplot as plt
# import pandas as pd
#
# rows = [
#     ('aa', 10, 130),
#     ('bb', 40, 160),
#     ('cc', 50, 170),
#     ('dd', 20, 110),
#     ('ee', 10, 140),
#     ('ff', 50, 130),
# ]
#
# df = pd.DataFrame(rows)
# df.columns = ['x','y','z']
# y=df.y
# z=df.z
#
# xrange=range(len(df.x))
# xlabel=df.x
#
#
# fig = plt.figure()
# ax1 = fig.add_subplot(111)
#
# ax1.bar(xrange,y,alpha = 0.3,color='blue',label=u'y')
#
# ax1.legend(loc=2)
# ax1.set_ylabel(u'z');
#
# ax2 = ax1.twinx()
# plt.plot(xrange, z,'or-',label=u'z');
# for i,(_x,_y) in enumerate(zip(xrange,z)):
#     plt.text(_x,_y,z[i],color='black',fontsize=10,)
# ax2.legend(loc=1)
#
# plt.xticks(xrange,xlabel)
#
# plt.show()
